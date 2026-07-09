"""excel_writer.py - Solo genera Excel nuevo con los datos OCR procesados."""
import logging
import math
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from src.config import OUTPUT_DIR, COLUMNAS

logger = logging.getLogger(__name__)


def _truncar_2_decimales(valor):
    """Trunca un valor a 2 decimales sin redondear."""
    if valor is None or valor == '':
        return valor
    try:
        num = float(valor)
        return math.trunc(num * 100) / 100
    except (ValueError, TypeError):
        return valor


def escribir_excel(registros: list, ruta_salida: str = None):
    """Crea un nuevo archivo Excel con los registros procesados."""
    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.title = "Coordenadas"

    # Escribir encabezados
    for col_idx, header in enumerate(COLUMNAS, 1):
        cell = hoja.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Escribir datos con formato de 2 decimales
    for i, reg in enumerate(registros):
        fila = i + 2  # Fila 2 en adelante
        for col_idx, campo in enumerate(COLUMNAS, 1):
            val = reg.get(campo, "")
            if col_idx > 1:  # X, Y, COTA, ABS (columnas numéricas)
                val = _truncar_2_decimales(val)
            cell = hoja.cell(row=fila, column=col_idx, value=val)
            if col_idx > 1:
                cell.number_format = '0.00'

    # Determinar ruta de salida
    if not ruta_salida:
        ruta_salida = OUTPUT_DIR / "resultado.xlsx"
    else:
        ruta_salida = Path(ruta_salida)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)
    logger.info(f"Excel guardado en: {ruta_salida}")
    return str(ruta_salida)
